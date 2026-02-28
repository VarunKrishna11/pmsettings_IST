#!/usr/bin/env python3
"""
generate_smelt_config.py - Generate SMELT Update Config
========================================================

Scans the SMELT_fitting folder structure, auto-matches each SMELT output
to ISTModeNames (from the base .pm file and/or Excel), and writes a
smelt_update_config.json that feeds into update_ist_coefficients.py.

Usage (run from pmsettings_IST/):
  python scripts/generate_smelt_config.py -o config/smelt_update_config.json

After generation, review the JSON:
  - Verify ist_modes[] for each smelt_entries[] item
  - Optionally add user_groups[] for modes that should share one VFE
  - Verify target configs and family

Then run:
  python scripts/update_ist_coefficients.py --config config/smelt_update_config.json
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
from typing import Dict, List

from ist_utils import (
    FAMILY_FOLDER_MAP,
    SUPPORTED_FAMILIES,
    detect_rail,
    get_family_smelt_root,
    parse_vfes,
)

# Project root: pmsettings_IST/ (one level above scripts/)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)

# Standard directories (resolved from script location, works from any CWD)
_INPUT_DIR = os.path.join(_PROJECT_ROOT, 'input')
_CONFIG_DIR = os.path.join(_PROJECT_ROOT, 'config')


def _in_input_dir(name: str) -> str:
    """Resolve a bare name relative to the project's input/ directory.
    Paths with separators or absolute paths are returned as-is.
    """
    if os.path.isabs(name) or os.sep in name or '/' in name:
        return name
    return os.path.join(_INPUT_DIR, name)


def _in_config_dir(name: str) -> str:
    """Resolve a bare name relative to the project's config/ directory.
    Paths with separators or absolute paths are returned as-is.
    """
    if os.path.isabs(name) or os.sep in name or '/' in name:
        return name
    return os.path.join(_CONFIG_DIR, name)


# ─────────────────────────────────────────────────────────────────
#  SMELT Folder -> ISTModeName Matching
# ─────────────────────────────────────────────────────────────────

def _tokens_from_folder(base: str) -> List[str]:
    """Extract classification tokens from a SMELT folder base name."""
    upper = base.upper()
    tokens = []
    for t in ['FTM', 'SDD', 'MBIST', 'PLL', 'RAMSEQ', 'CAD', 'CATA', 'SA', 'CAS',
              'BRIDGING', 'FSEQ', 'BASEFTM']:
        if t in upper:
            tokens.append(t)
    for t in ['P0H', 'P0M', 'P0L', 'P8']:
        if t in upper:
            tokens.append(t)
    for t in ['HT', 'LT']:
        if re.search(rf'_{t}(_|$)', upper):
            tokens.append(t)
    if 'P0L1800' in upper:
        if 'P0L' not in tokens:
            tokens.append('P0L')
    return tokens


def _score_mode(mode: str, tokens: List[str], folder_base: str = '') -> int:
    """Score how well an ISTModeName matches SMELT folder tokens."""
    upper = mode.upper()
    folder_upper = folder_base.upper()
    score = 0
    for t in tokens:
        if t in upper:
            score += 2
    for primary in ['MBIST', 'SDD', 'FTM', 'CAD', 'BRIDGING', 'FSEQ']:
        if primary in tokens and primary in upper:
            score += 2
    if 'VMIN' in folder_upper and 'VMAX' in upper:
        score -= 10
    if 'VMAX' in folder_upper and 'VMIN' in upper:
        score -= 10
    return score


def load_xlsx_mode_names(xlsx_path: str, families: List[str] = None) -> Dict[str, List[str]]:
    """
    Load ISTModeNames from IST_Modes_Support.xlsx.

    Returns dict: family -> list of mode names.
    The MATHS-IST sheet has a 'ModeName' column.
    The RIST sheet has a 'Description' column (col K, RISTItemEntry section).
    """
    if families is None:
        families = SUPPORTED_FAMILIES

    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed, cannot read xlsx. Using base file modes only.")
        return {}

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result: Dict[str, List[str]] = {}

    # MATHS-IST sheet
    if 'MATHS-IST' in families and 'MATHS-IST' in wb.sheetnames:
        ws = wb['MATHS-IST']
        header_row = None
        for r in range(1, 15):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if 'ModeName' in row:
                header_row = r
                break
        if header_row:
            headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
            col_map = {h: i + 1 for i, h in enumerate(headers) if h}
            if 'ModeName' in col_map:
                modes = []
                for r in range(header_row + 1, ws.max_row + 1):
                    val = ws.cell(r, col_map['ModeName']).value
                    if val:
                        modes.append(str(val).strip())
                result['MATHS-IST'] = list(dict.fromkeys(modes))

    # RIST sheet -- contains both RIST and RIST-Adaptive mode names
    if ('RIST' in families or 'RIST-Adaptive' in families) and 'RIST' in wb.sheetnames:
        ws = wb['RIST']
        rist_modes = []
        rist_adaptive_modes = []
        # The Description column (col K, index 10) in the RISTItemEntry section
        for r in range(3, ws.max_row + 1):
            desc = ws.cell(r, 11).value  # Column K = Description
            if desc:
                name = str(desc).strip()
                if name.endswith('_Adaptive'):
                    rist_adaptive_modes.append(name)
                else:
                    rist_modes.append(name)
        if 'RIST' in families:
            result['RIST'] = list(dict.fromkeys(rist_modes))
        if 'RIST-Adaptive' in families:
            result['RIST-Adaptive'] = list(dict.fromkeys(rist_adaptive_modes))

    return result


# ─────────────────────────────────────────────────────────────────
#  Config Generation
# ─────────────────────────────────────────────────────────────────

def generate_config(
    ist_base: str,
    xlsx_path: str,
    output_path: str,
    families: List[str] = None,
) -> int:
    """Generate smelt_update_config.json by scanning per-family SMELT folders."""

    if families is None:
        families = list(SUPPORTED_FAMILIES)

    # Collect ISTModeNames from all target families
    all_modes: List[str] = []
    modes_by_family: Dict[str, List[str]] = {}
    text = ''
    if ist_base and os.path.isfile(ist_base):
        with open(ist_base, 'r', encoding='utf-8') as f:
            text = f.read()
        for cfg_m in re.finditer(r"'(GR\d+-\w+)'\s*=>\s*\{", text):
            cfg_name = cfg_m.group(1)
            for fam in families:
                vfes = parse_vfes(text, cfg_name, family=fam)
                fam_modes = modes_by_family.setdefault(fam, [])
                for v in vfes:
                    for m in v['modes']:
                        if m not in fam_modes:
                            fam_modes.append(m)
                        if m not in all_modes:
                            all_modes.append(m)
        for fam in families:
            count = len(modes_by_family.get(fam, []))
            if count:
                print(f"  Found {count} {fam} modes from base file")
    else:
        print(f"  WARNING: ist_base not found: {ist_base}")

    if xlsx_path and os.path.isfile(xlsx_path):
        xlsx_modes_by_fam = load_xlsx_mode_names(xlsx_path, families)
        for fam, fam_xlsx_modes in xlsx_modes_by_fam.items():
            added = 0
            for m in fam_xlsx_modes:
                if m not in all_modes:
                    all_modes.append(m)
                    added += 1
            print(f"  Found {len(fam_xlsx_modes)} modes from xlsx {fam} (+{added} new)")

    print(f"  Total unique modes across families: {len(all_modes)}")

    # Scan per-family SMELT folders
    entries = []
    for fam in families:
        fam_smelt_root = get_family_smelt_root(_INPUT_DIR, fam)
        if not os.path.isdir(fam_smelt_root):
            print(f"  {fam}: no SMELT folder at {fam_smelt_root}, skipping")
            continue

        fam_modes = modes_by_family.get(fam, [])
        fam_entry_count = 0

        for item in sorted(os.listdir(fam_smelt_root)):
            item_path = os.path.join(fam_smelt_root, item)
            if not os.path.isdir(item_path):
                continue
            folder_base = re.sub(r'\.\d+$', '', item)
            rail = detect_rail(folder_base)

            coef_files = glob.glob(os.path.join(item_path, 'model.coef*.csv'))
            if not coef_files:
                continue

            # Auto-match against this family's modes only
            match_pool = fam_modes if fam_modes else all_modes
            tokens = _tokens_from_folder(folder_base)
            scored = [(m, _score_mode(m, tokens, folder_base)) for m in match_pool]
            scored = [(m, s) for m, s in scored if s > 0]
            scored.sort(key=lambda x: x[1], reverse=True)

            ist_modes = []
            candidates = []
            note = f"family: {fam} | rail: {rail or '?'}"

            if scored:
                best_score = scored[0][1]
                best = [m for m, s in scored if s == best_score]
                if len(best) == 1:
                    ist_modes = [best[0]]
                    note += f" | best match: {best[0]} (score={best_score})"
                else:
                    candidates = best[:6]
                    note += f" | AMBIGUOUS - pick one: {candidates}"

            entry = {
                'folder_base': folder_base,
                'family': fam,
                'ist_modes': ist_modes,
                'note': note,
            }
            if candidates:
                entry['_candidates'] = candidates
            entries.append(entry)
            fam_entry_count += 1

        print(f"  {fam}: scanned {fam_entry_count} SMELT folders from {FAMILY_FOLDER_MAP.get(fam, fam)}/SMELT_fitting/")

    # Detect target configs from base file
    target_configs = []
    if ist_base and os.path.isfile(ist_base):
        for m in re.finditer(r"'(GR\d+-\w+)'\s*=>\s*\{", text):
            target_configs.append(m.group(1))

    # Compute paths relative to the output config file's directory
    # (update_ist_coefficients.py resolves inputs relative to config dir)
    output_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(output_dir, exist_ok=True)

    def _relpath(p):
        """Make a path relative to the config file directory, using forward slashes."""
        return os.path.relpath(os.path.abspath(p), output_dir).replace('\\', '/')

    # Build per-family smelt_roots dict
    smelt_roots = {}
    for fam in families:
        fam_smelt_root = get_family_smelt_root(_INPUT_DIR, fam)
        if os.path.isdir(fam_smelt_root):
            smelt_roots[fam] = _relpath(fam_smelt_root)

    config = {
        '_description': 'Configuration for update_ist_coefficients.py',
        '_usage': [
            '1. Review smelt_entries[].ist_modes - fill in any empty/ambiguous ones',
            '2. Optionally add user_groups[] to merge modes into shared VFEs',
            '3. Run: python scripts/update_ist_coefficients.py --config <this_file>',
        ],
        'inputs': {
            'ist_base': _relpath(ist_base) if ist_base else 'ist_settings.pm',
            'smelt_roots': smelt_roots,
            'xlsx': _relpath(xlsx_path) if xlsx_path else 'IST_Modes_Support.xlsx',
        },
        'output': 'ist_settings.pm',
        'target': {
            'configs': target_configs,
            'families': families,
        },
        'smelt_entries': entries,
        '_user_groups_help': [
            "user_groups[] lets you merge multiple ISTModeNames into one shared VFE.",
            "By default (empty), each SMELT-covered mode gets its own VFE per rail.",
            "",
            "Scenario 1 - Group FTM and SDD at same pstate:",
            "  { \"ist_modes\": [\"BaseFTM2CLK_P0L\", \"SDD_P0L\"] }",
            "  -> Instead of 2 VFEs per rail, they share 1 VFE with same coefficients.",
            "",
            "Scenario 2 - Group all MBIST pstates together:",
            "  { \"ist_modes\": [\"MBIST_P0L\", \"MBIST_P0M_HT\", \"MBIST_P0H_HT\"] }",
            "  -> Instead of 3 MBIST VFEs per rail, they share 1 VFE per rail.",
            "",
            "Scenario 3 - Pull non-SMELT modes into a SMELT group:",
            "  { \"ist_modes\": [\"BaseFTM2CLK_P0L\", \"FseqRAMSeq_P0L\", \"Bridging_P0L\"] }",
            "  -> FseqRAMSeq_P0L, Bridging_P0L don't have SMELT data but get pulled",
            "     from residual VFE into BaseFTM2CLK_P0L's VFE, sharing its SMELT coefficients.",
        ],
        'user_groups': [],
    }

    # Write JSON config
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=2)

    # Write companion Excel config (same name, .xlsx extension)
    xlsx_output = os.path.splitext(output_path)[0] + '.xlsx'
    _export_config_to_xlsx(config, xlsx_output)

    print(f"\nGenerated config:")
    print(f"  JSON:  {output_path}")
    print(f"  Excel: {xlsx_output}")
    print(f"  SMELT entries: {len(entries)}")
    auto_ok = sum(1 for e in entries if e.get('ist_modes'))
    need_review = sum(1 for e in entries if not e.get('ist_modes'))
    print(f"  Auto-matched: {auto_ok}")
    print(f"  Need review:  {need_review}")
    if need_review:
        print("\n  Please review entries with empty ist_modes[] and fill them in.")
    print(f"\n  Edit either file, then run:")
    print(f"    python scripts/update_ist_coefficients.py --config {output_path}")
    print(f"    python scripts/update_ist_coefficients.py --config {xlsx_output}")
    return 0


def _export_config_to_xlsx(config: dict, xlsx_path: str) -> None:
    """Export config dict to a formatted Excel workbook with 3 sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("WARNING: openpyxl not installed, skipping Excel config export")
        return

    wb = openpyxl.Workbook()
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    need_review_fill = PatternFill("solid", fgColor="FCE4D6")  # orange for items needing review
    ok_fill = PatternFill("solid", fgColor="C6EFCE")           # green for auto-matched

    # ── Sheet 1: settings ──
    ws_settings = wb.active
    ws_settings.title = "settings"
    ws_settings.append(["Key", "Value"])
    for col in range(1, 3):
        c = ws_settings.cell(1, col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    inputs = config.get('inputs', {})
    ws_settings.append(["ist_base", inputs.get('ist_base', '')])
    # Per-family SMELT roots
    smelt_roots = inputs.get('smelt_roots', {})
    if smelt_roots:
        for fam, root in smelt_roots.items():
            ws_settings.append([f"smelt_root_{fam}", root])
    else:
        ws_settings.append(["smelt_root", inputs.get('smelt_root', '')])
    ws_settings.append(["xlsx", inputs.get('xlsx', '')])
    target = config.get('target', {})
    ws_settings.append(["target_configs", "|".join(target.get('configs', []))])
    target_families = target.get('families', [target.get('family', 'MATHS-IST')])
    if isinstance(target_families, str):
        target_families = [target_families]
    ws_settings.append(["target_families", "|".join(target_families)])

    ws_settings.column_dimensions['A'].width = 20
    ws_settings.column_dimensions['B'].width = 50
    ws_settings.freeze_panes = "A2"

    # ── Sheet 2: smelt_entries ──
    ws_entries = wb.create_sheet(title="smelt_entries")
    ws_entries.append(["folder_base", "family", "ist_modes", "note", "candidates"])
    for col in range(1, 6):
        c = ws_entries.cell(1, col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for entry in config.get('smelt_entries', []):
        ist_modes = "|".join(entry.get('ist_modes', []))
        candidates = "|".join(entry.get('_candidates', []))
        row_data = [
            entry.get('folder_base', ''),
            entry.get('family', ''),
            ist_modes,
            entry.get('note', ''),
            candidates,
        ]
        ws_entries.append(row_data)
        row_idx = ws_entries.max_row
        fill = ok_fill if ist_modes else need_review_fill
        for col in range(1, 6):
            ws_entries.cell(row_idx, col).fill = fill
            ws_entries.cell(row_idx, col).alignment = cell_align

    ws_entries.column_dimensions['A'].width = 45
    ws_entries.column_dimensions['B'].width = 16
    ws_entries.column_dimensions['C'].width = 30
    ws_entries.column_dimensions['D'].width = 60
    ws_entries.column_dimensions['E'].width = 40
    ws_entries.freeze_panes = "A2"

    # ── Sheet 3: user_groups ──
    ws_groups = wb.create_sheet(title="user_groups")
    ws_groups.append(["group_id", "ist_modes"])
    for col in range(1, 3):
        c = ws_groups.cell(1, col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for i, group in enumerate(config.get('user_groups', [])):
        ws_groups.append([f"group_{i}", "|".join(group.get('ist_modes', []))])

    ws_groups.column_dimensions['A'].width = 15
    ws_groups.column_dimensions['B'].width = 60
    ws_groups.freeze_panes = "A2"

    wb.save(xlsx_path)


# ─────────────────────────────────────────────────────────────────
#  CLI
# ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Generate SMELT update config by scanning per-family SMELT folders',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples (works from any directory):
  # Use all defaults (scans input/IST_MATHS/SMELT_fitting, etc.)
  python scripts/generate_smelt_config.py

  # Specific families only
  python scripts/generate_smelt_config.py --family MATHS-IST RIST

  # Custom output name (just the name -- automatically goes to config/)
  python scripts/generate_smelt_config.py -o my_config.json

Per-family SMELT folders are auto-discovered under input/:
  input/IST_MATHS/SMELT_fitting/         -> MATHS-IST
  input/IST_RIST/SMELT_fitting/           -> RIST
  input/IST_RIST_Adaptive/SMELT_fitting/  -> RIST-Adaptive

NOTE: Bare names (no path separators) are resolved relative to the project's
      input/ directory (for inputs) or config/ directory (for -o output).
""")

    parser.add_argument('--ist-base', type=str, default='ist_settings.pm',
                        help='Base .pm file name or path (default: ist_settings.pm -> input/ist_settings.pm)')
    parser.add_argument('--xlsx', type=str, default='IST_Modes_Support.xlsx',
                        help='Excel modes file name or path (default: IST_Modes_Support.xlsx -> input/...)')
    parser.add_argument('-o', '--output', type=str, default='smelt_update_config.json',
                        help='Output config name or path (default: smelt_update_config.json -> config/...)')
    parser.add_argument('--family', type=str, nargs='+',
                        default=None,
                        help='Target families (default: all). '
                             'Options: MATHS-IST RIST RIST-Adaptive')

    args = parser.parse_args()

    families = args.family
    if families:
        for f in families:
            if f not in SUPPORTED_FAMILIES:
                print(f"ERROR: Unknown family '{f}'. Supported: {SUPPORTED_FAMILIES}")
                return 1

    return generate_config(
        ist_base=_in_input_dir(args.ist_base),
        xlsx_path=_in_input_dir(args.xlsx),
        output_path=_in_config_dir(args.output),
        families=families,
    )


if __name__ == '__main__':
    sys.exit(main())
